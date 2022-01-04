from openpyxl.reader import excel
import requests
from bs4 import BeautifulSoup
import codecs
import openpyxl
from openpyxl import Workbook

path = "C:/Users/Mr. Paul/Downloads/商家中心－-理想生活上天猫.html"
excel_path = "C:/Users/Mr. Paul/Desktop/ItemDescription.xlsx"
f=codecs.open("C:/Users/Mr. Paul/Downloads/商家中心－-理想生活上天猫.html", 'r', encoding='utf-8')
#print(f.read())

id_element = []
description_element = []
status_element = []
price_element = []
old_id_element = []
sale_element = []
create_new_sheet = True

soup = BeautifulSoup(f, "html.parser")
results = soup.find(id="root")
# print(results.prettify())
job_elements = soup.find_all("tr", class_="next-table-row")
group_elements = soup.find_all("div", class_="queryShopCategoryId-sell-hoc")
group_element = group_elements[0].text.strip()[4:]
for job_element in job_elements:
    parent_elements = job_element.find_all("span", class_="product-desc-span")
    id_element.append(parent_elements[1].text.strip()[3:])
    description_element.append(parent_elements[0].text.strip())
    if(len(parent_elements) == 5):
        status_element.append(parent_elements[4].text.strip())
        old_id_element.append('N/A')
        price_element.append(parent_elements[2].text.strip())
    else:
        status_element.append(parent_elements[5].text.strip())
        old_id_element.append(parent_elements[2].text.strip()[3:])
        price_element.append(parent_elements[3].text.strip())
    sale_elements = job_element.find_all("span", class_="table-text-cell")
    sale_element.append(sale_elements[1].text.strip())

data = [id_element, 
        description_element,
        old_id_element,
        price_element,
        sale_element,
        status_element,
        ]

wb_obj = openpyxl.load_workbook(r'C:\Users\Mr. Paul\Desktop\ItemDescription.xlsx')
for sheetname in wb_obj.sheetnames:
    # print(sheetname, group_element)
    if sheetname == group_element:
        create_new_sheet = False
# print(create_new_sheet)
header_name = ['id','description','package','old_id','new_id','off stock','price','sale','status']
if create_new_sheet:
    ws = wb_obj.create_sheet(group_element)
    ws = wb_obj[group_element]
    for row in ws.iter_rows(min_row=1, max_col=9, max_row=1):
        row = list(row)
        row[0].value = header_name[0]
        row[1].value = header_name[1]
        row[2].value = header_name[2]
        row[3].value = header_name[3]
        row[4].value = header_name[4]
        row[5].value = header_name[5]
        row[6].value = header_name[6]
        row[7].value = header_name[7]
        row[8].value = header_name[8]
else:
    ws = wb_obj[group_element]
new_row = ws.max_row + 1
total_inserted_rows = len(id_element)
counter = 0

for row in ws.iter_rows(min_row=new_row, max_col=9, max_row=(new_row + total_inserted_rows-1)):
    row = list(row)
    #print(len(row),row[8].value, data[0][19],counter)
    row[0].value = data[0][counter]
    row[1].value = data[1][counter]
    row[3].value = data[2][counter]
    row[6].value = data[3][counter]
    row[7].value = data[4][counter]
    row[8].value = data[5][counter]
    counter += 1

openpyxl.Workbook.save(wb_obj, excel_path)